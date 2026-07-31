from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from finance_agent.voucher_record import SourceCell, VoucherRecord, decimal_to_money


STANDARD_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "voucher_date": ("日期", "凭证日期", "付款日期", "交易日期", "回单日期"),
    "voucher_no": ("凭证编号", "凭证号", "凭证字号", "记账凭证号"),
    "summary": ("摘要", "用途", "事由", "付款摘要"),
    "income_amount": ("收入/拨入", "收入", "拨入", "借方金额"),
    "expense_amount": ("支出/拨出", "支出", "拨出", "贷方金额", "付款金额", "金额"),
    "budget_category": ("预算分类", "预算科目", "经费类别", "支出类别"),
    "attachment_note": ("附件材料", "附件", "备注", "项目经费号", "经费号"),
}

BUDGET_CATEGORY_RE = re.compile(r"^[（(]?\s*(?P<code>\d+)\s*[）)]?\s*(?P<name>.+)$")
PROJECT_FUND_RE = re.compile(r"\b\d{5,}-\d{5,}\b")
LEADING_PROJECT_CODE_RE = re.compile(r"^\s*(?P<code>\d{5,8})\b")
REFERENCE_RE = re.compile(r"\b(?:20\d{6,}|YQJYJJBZD\d{8,}|APC\d+|1-\d{8,})\b", re.I)


RECORD_ID_RULE = "sha1(source_file_name|sheet_name|excel_row|voucher_no|voucher_date)[:16]"


def parse_workbook(
    input_path: Path,
    raw_archive_dir: Path | None = None,
    include_debug: bool = False,
) -> dict[str, Any]:
    input_path = input_path.resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    archived_source = archive_source_file(input_path, raw_archive_dir) if raw_archive_dir else None
    workbook = load_workbook(input_path, data_only=True, read_only=False)

    sheet_reports: list[dict[str, Any]] = []
    records: list[VoucherRecord] = []
    for sheet in workbook.worksheets:
        report, sheet_records = parse_sheet(sheet, input_path)
        sheet_reports.append(report)
        records.extend(sheet_records)

    payload: dict[str, Any] = {
        "schema": "VoucherRecord",
        "schema_version": "2.0",
        "purpose": "research_project_financial_acceptance_analysis",
        "record_id_rule": RECORD_ID_RULE,
        "record_count": len(records),
        "records": [record.to_agent_dict() for record in records],
    }
    if include_debug:
        payload["debug"] = {
            "source_file": str(input_path),
            "archived_source_file": str(archived_source) if archived_source else None,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "mapping_report": sheet_reports,
            "records": [record.to_debug_dict() for record in records],
        }
    return payload


def parse_sheet(sheet: Worksheet, source_file: Path) -> tuple[dict[str, Any], list[VoucherRecord]]:
    header_row_index = detect_header_row(sheet)
    headers = [normalize_header(cell.value) for cell in sheet[header_row_index]]
    field_map = map_headers(headers, sheet, header_row_index)

    records: list[VoucherRecord] = []
    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        cells = list(sheet[row_index])
        if is_blank_row(cell.value for cell in cells):
            continue

        raw_by_header = build_raw_row(headers, cells)
        warnings: list[str] = []

        voucher_date = normalize_date(get_field_value(field_map, "voucher_date", cells), warnings)
        voucher_no = normalize_text(get_field_value(field_map, "voucher_no", cells))
        summary = normalize_text(get_field_value(field_map, "summary", cells))
        income = parse_decimal(get_field_value(field_map, "income_amount", cells), "income_amount", warnings)
        expense = parse_decimal(get_field_value(field_map, "expense_amount", cells), "expense_amount", warnings)
        direction, signed_amount = derive_direction_and_amount(income, expense, warnings)
        budget_code, budget_name = parse_budget_category(get_field_value(field_map, "budget_category", cells))
        attachment_note = normalize_text(get_field_value(field_map, "attachment_note", cells))
        fund_owner, project_fund_no = parse_attachment_note(attachment_note)
        project_code = parse_project_code(summary)
        references = parse_reference_numbers(summary)

        record = VoucherRecord(
            record_id=make_record_id(source_file, sheet.title, row_index, voucher_no, voucher_date),
            source_file=str(source_file),
            source_sheet=sheet.title,
            source_row=row_index,
            voucher_date=voucher_date,
            voucher_no=voucher_no,
            summary=summary,
            income_amount=decimal_to_money(income),
            expense_amount=decimal_to_money(expense),
            signed_amount=decimal_to_money(signed_amount),
            direction=direction,
            budget_category_code=budget_code,
            budget_category_name=budget_name,
            project_code=project_code,
            project_fund_no=project_fund_no,
            fund_owner=fund_owner,
            bank_receipt_reference_numbers=references,
            attachment_note=attachment_note,
            source_cells=build_source_cells(sheet.title, headers, cells, field_map),
            raw_row=raw_by_header,
            warnings=warnings,
        )
        records.append(record)

    report = {
        "sheet": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "header_row": header_row_index,
        "headers": headers,
        "field_map": {field: {"column": idx + 1, "header": headers[idx]} for field, idx in field_map.items()},
        "record_count": len(records),
    }
    return report, records


def detect_header_row(sheet: Worksheet, max_scan_rows: int = 10) -> int:
    best_row = 1
    best_score = -1
    aliases = {alias for names in STANDARD_FIELD_ALIASES.values() for alias in names}
    for row_index in range(1, min(sheet.max_row, max_scan_rows) + 1):
        values = [normalize_header(cell.value) for cell in sheet[row_index]]
        score = sum(1 for value in values if value in aliases)
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


def map_headers(headers: list[str | None], sheet: Worksheet, header_row_index: int) -> dict[str, int]:
    normalized = {header: idx for idx, header in enumerate(headers) if header}
    mapping: dict[str, int] = {}
    for standard_field, aliases in STANDARD_FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[standard_field] = normalized[alias]
                break

    for idx, header in enumerate(headers):
        if header is None and looks_like_attachment_note_column(sheet, idx + 1, header_row_index):
            mapping.setdefault("attachment_note", idx)
            break
    return mapping


def looks_like_attachment_note_column(sheet: Worksheet, column_index: int, header_row_index: int) -> bool:
    matches = 0
    non_blank = 0
    for row_index in range(header_row_index + 1, min(sheet.max_row, header_row_index + 15) + 1):
        value = normalize_text(sheet.cell(row=row_index, column=column_index).value)
        if not value:
            continue
        non_blank += 1
        if PROJECT_FUND_RE.search(value):
            matches += 1
    return non_blank > 0 and matches / non_blank >= 0.5


def build_raw_row(headers: list[str | None], cells: list[Any]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for idx, cell in enumerate(cells):
        header = headers[idx] if idx < len(headers) else None
        key = header or f"unnamed_column_{idx + 1}"
        raw[key] = normalize_json_value(cell.value)
    return raw


def build_source_cells(
    sheet_name: str,
    headers: list[str | None],
    cells: list[Any],
    field_map: dict[str, int],
) -> dict[str, SourceCell]:
    sources: dict[str, SourceCell] = {}
    for field, idx in field_map.items():
        if idx >= len(cells):
            continue
        cell = cells[idx]
        sources[field] = SourceCell(
            sheet=sheet_name,
            row=cell.row,
            column=cell.column,
            header=headers[idx] if idx < len(headers) else None,
            raw_value=normalize_json_value(cell.value),
        )
    return sources


def derive_direction_and_amount(
    income: Decimal | None,
    expense: Decimal | None,
    warnings: list[str],
) -> tuple[str | None, Decimal | None]:
    has_income = income is not None and income != 0
    has_expense = expense is not None and expense != 0
    if has_income and has_expense:
        warnings.append("income_amount and expense_amount are both non-zero")
    if has_expense:
        return "expense", -expense
    if has_income:
        return "income", income
    warnings.append("no non-zero amount found")
    return None, None


def normalize_date(value: Any, warnings: list[str]) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    warnings.append(f"unrecognized date format: {text}")
    return text


def parse_decimal(value: Any, field_name: str, warnings: list[str]) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = normalize_text(value)
    if not text:
        return None
    cleaned = text.replace(",", "").replace("￥", "").replace("¥", "").strip()
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        warnings.append(f"unrecognized decimal in {field_name}: {text}")
        return None


def parse_budget_category(value: Any) -> tuple[str | None, str | None]:
    text = normalize_text(value)
    if not text:
        return None, None
    match = BUDGET_CATEGORY_RE.match(text)
    if match:
        return match.group("code"), match.group("name").strip()
    return None, text


def parse_attachment_note(value: str | None) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    fund_match = PROJECT_FUND_RE.search(value)
    fund_no = fund_match.group(0) if fund_match else None
    applicant = value
    if fund_no:
        applicant = value.replace(fund_no, "").strip()
    return applicant or None, fund_no


def parse_project_code(summary: str | None) -> str | None:
    if not summary:
        return None
    match = LEADING_PROJECT_CODE_RE.match(summary)
    return match.group("code") if match else None


def parse_reference_numbers(summary: str | None) -> list[str]:
    if not summary:
        return []
    return sorted(set(match.group(0) for match in REFERENCE_RE.finditer(summary)))


def archive_source_file(input_path: Path, raw_archive_dir: Path) -> Path:
    raw_archive_dir.mkdir(parents=True, exist_ok=True)
    target = raw_archive_dir / input_path.name
    if target.resolve() == input_path.resolve():
        return target
    if target.exists():
        digest = hashlib.sha256(input_path.read_bytes()).hexdigest()[:10]
        target = raw_archive_dir / f"{input_path.stem}.{digest}{input_path.suffix}"
    shutil.copy2(input_path, target)
    return target


def write_json(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def get_field_value(field_map: dict[str, int], field: str, cells: list[Any]) -> Any:
    idx = field_map.get(field)
    if idx is None or idx >= len(cells):
        return None
    return cells[idx].value


def normalize_header(value: Any) -> str | None:
    text = normalize_text(value)
    return text if text else None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\u3000", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def normalize_json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    return value


def is_blank_row(values: Iterable[Any]) -> bool:
    return all(normalize_text(value) is None for value in values)


def make_record_id(source_file: Path, sheet_name: str, row_index: int, voucher_no: str | None, voucher_date: str | None) -> str:
    base = f"{source_file.name}|{sheet_name}|{row_index}|{voucher_no or ''}|{voucher_date or ''}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Convert voucher bank receipt checklist Excel files to VoucherRecord JSON.")
    parser.add_argument("input", type=Path, help="Path to the source .xlsx file.")
    parser.add_argument("-o", "--output", type=Path, default=Path("data/processed/voucher_records.json"))
    parser.add_argument("--archive-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--include-debug", action="store_true", help="Include source cell, raw row, and mapping report details.")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    payload = parse_workbook(args.input, raw_archive_dir=args.archive_dir, include_debug=args.include_debug)
    write_json(payload, args.output)
    print(f"Wrote {payload['record_count']} VoucherRecord rows to {args.output}")


if __name__ == "__main__":
    main()
